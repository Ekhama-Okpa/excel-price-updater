
import openpyxl, os
os.chdir("C:\\Users\\Owner\\Downloads")
wb = openpyxl.load_workbook("example1.xlsx")

sheet2 = wb["Sheet2"]

# Updated produce price dictionary.
updated_produce = {"Celery": 1.19, "Garlic": 3.07, "Lemon": 1.27}

# Loop from the second to last row in the sheet.
for row in range(2, sheet2.max_row + 1):
    # Checking if produce in the column "A" is the same as any one present in the dictonary.
    if sheet2[f"A{row}"].value in updated_produce.keys():
        # Replaces price of produce with corresponding value present in the dictionary
        sheet2[f"B{row}"].value = updated_produce[sheet2[f"A{row}"].value]

""" This section adds new produce present in the dictionary if not already in the sheet. """

# Creates a list of the produce present.
list_items = [item[0].value for item in sheet2[f"A2": f"A{sheet2.max_row}"]]

# Loops through the produce in the dictionary.
for produce in updated_produce.keys():
    value = sheet2.max_row + 1
    # Checks if the produce is not in the sheet.
    if produce not in list_items:
        # Adds a new produce to the bottom of the sheet alongside it's price and a formula for the total.
        # The "POUNDS SOLD" can be updated at any time.
        sheet2[f"A{value}"], sheet2[f"B{value}"], sheet2[f"D{value}"] = \
            produce, updated_produce[produce], f"=ROUND(B{value}*C{value}, 2)"

wb.save("example1.xlsx")
